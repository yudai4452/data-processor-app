import os
import pandas as pd
import openpyxl
import pytz
from openpyxl.styles import PatternFill, Font
from bs4 import BeautifulSoup
import streamlit as st
from datetime import datetime, timedelta, timezone
from github import Github
import plotly.graph_objects as go 

# ダウンロード用にファイルを読み込む関数
def download_excel_file(excel_path):
    with open(excel_path, "rb") as f:
        return f.read()

# GitHubへのファイルアップロード関数
def upload_file_to_github(file_path, repo_name, file_name_in_repo, commit_message):
    try:
        g = Github(GITHUB_TOKEN)
        repo = g.get_repo(repo_name)

        with open(file_path, 'rb') as file:
            content = file.read()

        path = file_name_in_repo

        try:
            contents = repo.get_contents(path)
            repo.update_file(path, commit_message, content, contents.sha)
            st.info(f"{file_name_in_repo} を更新しました。")
        except Exception as e_inner:
            repo.create_file(path, commit_message, content)
            st.info(f"{file_name_in_repo} を作成しました。")
    except Exception as e_outer:
        st.error(f"GitHubへのファイルアップロード中にエラーが発生しました: {e_outer}")
        st.error(f"詳細: {e_outer.args}")

def extract_data_and_save_to_csv(html_path, output_csv_path, date):
    with open(html_path, "r", encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "lxml")
    rows = soup.find_all("tr")[1:]

    data = {
        "台番号": [], "累計スタート": [], "BB回数": [], "RB回数": [], 
        "ART回数": [], "最大持玉": [], "BB確率": [], "RB確率": [], 
        "ART確率": [], "合成確率": []
    }

    for row in rows:
        cells = row.find_all("td")
        if len(cells) > 1:
            data["台番号"].append(cells[1].get_text())
            data["累計スタート"].append(cells[2].get_text())
            data["BB回数"].append(cells[3].get_text())
            data["RB回数"].append(cells[4].get_text())
            data["ART回数"].append(cells[5].get_text())
            data["最大持玉"].append(cells[6].get_text())
            data["BB確率"].append(cells[7].get_text())
            data["RB確率"].append(cells[8].get_text())
            data["ART確率"].append(cells[9].get_text())
            data["合成確率"].append(cells[10].get_text())

    df = pd.DataFrame(data)
    df.to_csv(output_csv_path, index=False, encoding="shift-jis")
    return df

def load_excel_data(excel_path):
    df = pd.read_excel(excel_path, sheet_name="合成確率", engine="openpyxl", index_col=0)
    return df

def plot_synthetic_probabilities(df, selected_machine_number):
    machine_data = df.loc[selected_machine_number].dropna()
    dates = machine_data.index
    probabilities = machine_data.values

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=dates, y=probabilities, mode='lines+markers', name=f'合成確率: {selected_machine_number}'))
    fig.update_layout(
        title=f"台番号 {selected_machine_number} の合成確率の推移",
        xaxis_title="日付",
        yaxis_title="合成確率",
        xaxis=dict(tickformat="%Y-%m-%d"),
        hovermode="x"
    )
    st.plotly_chart(fig)

def create_new_excel_with_all_data(output_csv_dir, excel_path):
    csv_files = [os.path.join(output_csv_dir, f) for f in os.listdir(output_csv_dir) if f.endswith('.csv')]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合成確率"

    ws.cell(row=1, column=1, value="台番号")
    
    all_data = {}
    date_columns = []

    for csv_file in csv_files:
        df = pd.read_csv(csv_file, encoding="shift-jis")
        date = os.path.basename(csv_file).split('_')[-1].replace('.csv', '')
        formatted_date = pd.to_datetime(date).strftime('%Y/%m/%d')
        date_columns.append(formatted_date)
        
        for index, row in df.iterrows():
            if row['台番号'] not in all_data:
                all_data[row['台番号']] = {}
            all_data[row['台番号']][formatted_date] = row['合成確率']
    
    for col_index, date in enumerate(sorted(date_columns), start=2):
        ws.cell(row=1, column=col_index, value=date)

    for row_index, (machine_number, dates_data) in enumerate(all_data.items(), start=2):
        ws.cell(row=row_index, column=1, value=machine_number)
        for col_index, date in enumerate(sorted(date_columns), start=2):
            ws.cell(row=row_index, column=col_index, value=dates_data.get(date, None))

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max(max_length + 2, 10)
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 20

    mei_font = Font(name="メイリオ")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = mei_font

    wb.save(excel_path)

def apply_color_fill_to_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            try:
                cell_value = float(cell.value)
                if cell_value < 125:
                    cell.fill = yellow_fill
                elif 125 <= cell_value < 140:
                    cell.fill = light_blue_fill
            except (TypeError, ValueError):
                pass

    wb.save(excel_path)

def process_juggler_data(html_path, output_csv_dir, excel_path, date):
    output_csv_path = os.path.join(output_csv_dir, f"slot_machine_data_{date}.csv")
    df_new = extract_data_and_save_to_csv(html_path, output_csv_path, date)

    create_new_excel_with_all_data(output_csv_dir, excel_path)
    apply_color_fill_to_excel(excel_path)

    print(f"データ処理が完了し、{excel_path} に保存されました")

# CSVとExcelファイルを両方ダウンロード可能にする
def display_download_buttons(output_csv_path, excel_path):
    if os.path.exists(output_csv_path):
        with open(output_csv_path, "rb") as f:
            st.download_button(
                label="CSVファイルをダウンロード",
                data=f,
                file_name=os.path.basename(output_csv_path),
                mime="text/csv",
                key="csv_download_button"
            )
    if os.path.exists(excel_path):
        with open(excel_path, "rb") as f:
            st.download_button(
                label="Excelファイルをダウンロード",
                data=f,
                file_name=excel_path,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="excel_download_button"
            )

# シークレットからGitHubトークンを取得
GITHUB_TOKEN = st.secrets["github"]["token"]

# 日本時間の今日の日付を取得
japan_time_zone = pytz.timezone('Asia/Tokyo')
current_date_japan = datetime.now(japan_time_zone)

st.markdown(
    """
    <style>
    .main-title {font-size: 40px; font-weight: bold; color: #34495E; text-align: center; margin-bottom: 20px;}
    .subtitle {font-size: 20px; color: #2ECC71; text-align: center; margin-bottom: 30px;}
    .section-title {font-size: 26px; font-weight: bold; color: #2980B9; margin-top: 30px; margin-bottom: 10px;}
    .instruction {font-size: 18px; color: #2C3E50; margin-bottom: 20px;}
    </style>
    <div class="main-title">🎰 Juggler Data Manager 🎰</div>
    <div class="subtitle">HTMLファイルからデータを抽出し、Excelファイルを生成します。</div>
    """, unsafe_allow_html=True
)

st.markdown(
    """
    <div class="section-title">📊 メッセ武蔵境店 台データオンライン</div>
    <div class="instruction">
        台データはこちらのリンクからご確認ください： 
        <a href="https://daidata.goraggio.com/100686" target="_blank" style="color: #3498DB; text-decoration: none;">メッセ武蔵境店 - 台データオンライン</a>
    </div>
    """, unsafe_allow_html=True
)

st.markdown(
    """
    <div class="section-title">🔍 HTMLの取得方法</div>
    <div class="instruction">
        1. リンク先のページにアクセス<br>
        2. 右上にあるメニューを押し、「その他」を選択<br>
        3. 「デベロッパーツール」を選択<br>
        4. 画面左上の1行目に表示される「&lt;html&gt;」を右クリック<br>
        5. 「copy」を選択し、「copy element」をクリック<br>
    </div>
    """, unsafe_allow_html=True
)

st.markdown(
    """
    <div class="section-title">📋 HTMLを貼り付ける方法</div>
    <div class="instruction">
        HTMLを以下のフォームに貼り付けたら、<strong>Ctrl + Enter</strong>を押して処理を<br>開始してください。
    </div>
    """, unsafe_allow_html=True
)

st.sidebar.markdown(
    """
    <style>
    .sidebar-title {font-size: 22px; font-weight: bold; color: #2980B9; margin-bottom: 15px;}
    .sidebar-section {font-size: 18px; color: #34495E; margin-bottom: 10px;}
    </style>
    """, unsafe_allow_html=True
)

st.sidebar.markdown('<div class="sidebar-title">📋 入力パラメータ</div>', unsafe_allow_html=True)
st.sidebar.markdown('<div class="sidebar-section">HTMLファイルの入力方法を選択してください</div>', unsafe_allow_html=True)

input_option = st.sidebar.radio("HTMLの入力方法を選択", ('ファイルをアップロード', 'HTMLを貼り付け'), index=1)

if input_option == 'ファイルをアップロード':
    st.sidebar.markdown('<div class="sidebar-section">HTMLファイルをアップロードしてください。</div>', unsafe_allow_html=True)
    uploaded_html = st.sidebar.file_uploader("HTMLファイルをアップロード", type=["html", "htm", "txt"])
    html_content = None
else:
    st.sidebar.markdown('<div class="sidebar-section">HTMLを貼り付けてください。貼り付け後に Ctrl + Enter を押してください。</div>', unsafe_allow_html=True)
    html_content = st.sidebar.text_area("HTMLを貼り付け", height=300)
    uploaded_html = None

st.sidebar.text_input("CSVファイルの保存フォルダ名", "マイジャグラーV", disabled=True)
excel_file_name = st.sidebar.text_input("Excelファイル名", "マイジャグラーV_塗りつぶし済み.xlsx")
date_input = st.sidebar.date_input("日付を選択", current_date_japan)
confirm_date = st.sidebar.checkbox(f"選択した日付は {date_input} です。確認しましたか？")

if os.path.exists(excel_file_name):
    df_synthetic = load_excel_data(excel_file_name)
    machine_numbers = df_synthetic.index.tolist()
    selected_machine_number = st.sidebar.selectbox("台番号を選択", machine_numbers)
    if selected_machine_number:
        plot_synthetic_probabilities(df_synthetic, selected_machine_number)

if st.sidebar.button("処理開始"):
    if confirm_date:
        if uploaded_html or html_content:
            if uploaded_html is not None:
                html_path = os.path.join(".", uploaded_html.name)
                with open(html_path, "wb") as f:
                    f.write(uploaded_html.getbuffer())
            else:
                html_path = os.path.join(".", "uploaded_html.html")
                with open(html_path, "w", encoding="utf-8") as f:
                    f.write(html_content)

            if not os.path.exists("マイジャグラーV"):
                os.makedirs("マイジャグラーV")

            date_str = date_input.strftime("%Y-%m-%d")

            try:
                process_juggler_data(html_path, "マイジャグラーV", excel_file_name, date_str)
                st.success(f"データ処理が完了し、{excel_file_name} に保存されました。")

                output_csv_path = os.path.join("マイジャグラーV", f"slot_machine_data_{date_str}.csv")
                display_download_buttons(output_csv_path, excel_file_name)

                repo_name = "yudai4452/data-processor-app"
                commit_message = f"Add data for {date_str}"

                upload_file_to_github(output_csv_path, repo_name, f"マイジャグラーV/slot_machine_data_{date_str}.csv", commit_message)
                upload_file_to_github(excel_file_name, repo_name, f"{excel_file_name}", commit_message)

            except Exception as e:
                st.error(f"エラーが発生しました: {e}")
        else:
            st.warning("HTMLファイルをアップロードするか、HTMLを貼り付けてください。")
    else:
        st.warning("日付の確認を行ってください。")
